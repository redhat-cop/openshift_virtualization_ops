#!/usr/bin/python
# GNU General Public License v3.0+ (see LICENSE or https://www.gnu.org/licenses/gpl-3.0.txt)

"""Ansible module to parse RVTools VMware inventory exports."""

from __future__ import absolute_import, division, print_function

__metaclass__ = type

DOCUMENTATION = r"""
---
module: rvtools_parse
short_description: Parse RVTools VMware inventory exports
version_added: "1.1.0"
description:
  - Reads an RVTools export file (xlsx or csv) and returns normalized,
    structured data for each supported tab.
  - Supports both xlsx (multi-tab) and csv (single-tab per file) formats.
options:
  src:
    description:
      - Path to the RVTools export file.
      - For xlsx format, this is a single file containing all tabs.
      - For csv format, this is one file per tab; use C(src_map) for multiple files.
    type: path
    required: false
  src_map:
    description:
      - Dictionary mapping tab names to CSV file paths.
      - Only used when C(format=csv).
      - "Keys: vInfo, vCPU, vMemory, vDisk, vNetwork, vSnapshot."
    type: dict
    required: false
  format:
    description:
      - Export file format.
    type: str
    choices: [xlsx, csv]
    required: true
  tabs:
    description:
      - List of tabs to parse. If omitted, all supported tabs are parsed.
    type: list
    elements: str
    required: false
    default: []
  encoding:
    description:
      - Character encoding for CSV files.
    type: str
    default: utf-8-sig
author:
  - OpenShift Virtualization Migration Contributors
"""

EXAMPLES = r"""
- name: Parse RVTools xlsx export
  infra.openshift_virtualization_ops.rvtools_parse:
    src: /data/rvtools_export.xlsx
    format: xlsx
  register: rvtools_data

- name: Parse only vInfo and vCPU tabs
  infra.openshift_virtualization_ops.rvtools_parse:
    src: /data/rvtools_export.xlsx
    format: xlsx
    tabs:
      - vInfo
      - vCPU
  register: rvtools_data

- name: Parse RVTools CSV exports
  infra.openshift_virtualization_ops.rvtools_parse:
    format: csv
    src_map:
      vInfo: /data/rvtools_vinfo.csv
      vCPU: /data/rvtools_vcpu.csv
      vNetwork: /data/rvtools_vnetwork.csv
  register: rvtools_data
"""

RETURN = r"""
parsed:
  description: Parsed RVTools data organized by tab name.
  returned: success
  type: dict
  contains:
    vinfo:
      description: VM general information.
      type: list
      elements: dict
    vcpu:
      description: VM CPU configuration.
      type: list
      elements: dict
    vmemory:
      description: VM memory configuration.
      type: list
      elements: dict
    vdisk:
      description: VM disk configuration.
      type: list
      elements: dict
    vnetwork:
      description: VM network configuration.
      type: list
      elements: dict
    vsnapshot:
      description: VM snapshot information.
      type: list
      elements: dict
tab_counts:
  description: Number of records parsed per tab.
  returned: success
  type: dict
  sample: {"vinfo": 150, "vcpu": 150, "vdisk": 320}
"""

import csv
import os
import traceback

from ansible.module_utils.basic import AnsibleModule

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    OPENPYXL_IMPORT_ERROR = traceback.format_exc()

TAB_FIELD_MAP = {
    "vInfo": {
        "VM": "vm_name",
        "Powerstate": "power_state",
        "Guest OS": "guest_os",
        "CPUs": "cpus",
        "Memory": "memory_mb",
        "NICs": "nics",
        "Disks": "disks",
        "Folder": "folder",
        "Resource pool": "resource_pool",
        "Cluster": "cluster",
        "Host": "host",
        "Datacenter": "datacenter",
        "DNS Name": "dns_name",
        "Primary IP Address": "primary_ip",
        "Annotation": "annotation",
    },
    "vCPU": {
        "VM": "vm_name",
        "CPUs": "cpus",
        "Sockets": "sockets",
        "Cores per Socket": "cores_per_socket",
        "CPU reservation [MHz]": "cpu_reservation_mhz",
        "CPU limit [MHz]": "cpu_limit_mhz",
    },
    "vMemory": {
        "VM": "vm_name",
        "Size [MB]": "size_mb",
        "Reservation [MB]": "reservation_mb",
        "Limit [MB]": "limit_mb",
        "Shares": "shares",
    },
    "vDisk": {
        "VM": "vm_name",
        "Disk": "disk_name",
        "Capacity MB": "capacity_mb",
        "Thin": "thin",
        "Datastore": "datastore",
        "Path": "path",
    },
    "vNetwork": {
        "VM": "vm_name",
        "Network Adapter": "nic_name",
        "Network": "network_name",
        "Adapter Type": "adapter_type",
        "MAC Address": "mac_address",
        "IP Address": "ip_address",
        "Connected": "connected",
    },
    "vSnapshot": {
        "VM": "vm_name",
        "Name": "snapshot_name",
        "Date / time": "created",
        "Description": "description",
        "Size [MB]": "size_mb",
    },
}

BOOL_FIELDS = {"thin", "connected"}
INT_FIELDS = {
    "cpus", "memory_mb", "nics", "disks", "sockets", "cores_per_socket",
    "cpu_reservation_mhz", "cpu_limit_mhz", "size_mb", "reservation_mb",
    "limit_mb", "shares", "capacity_mb",
}


def _coerce_value(field_name, value):
    """Coerce a parsed value to the correct Python type."""
    if value is None:
        return None
    if field_name in BOOL_FIELDS:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("true", "1", "yes")
    if field_name in INT_FIELDS:
        if isinstance(value, (int, float)):
            return int(value)
        try:
            cleaned = str(value).strip().replace(",", "")
            return int(float(cleaned)) if cleaned else 0
        except (ValueError, TypeError):
            return 0
    if isinstance(value, str):
        return value.strip()
    return value


def _parse_sheet(headers, rows, tab_name):
    """Parse rows from a sheet using the field map for the given tab."""
    field_map = TAB_FIELD_MAP.get(tab_name, {})
    col_indices = {}
    for idx, header in enumerate(headers):
        header_str = str(header).strip() if header else ""
        if header_str in field_map:
            col_indices[idx] = field_map[header_str]

    records = []
    for row in rows:
        record = {}
        for idx, norm_name in col_indices.items():
            value = row[idx] if idx < len(row) else None
            record[norm_name] = _coerce_value(norm_name, value)
        if record.get("vm_name"):
            records.append(record)
    return records


def _parse_xlsx(src, tabs):
    """Parse an RVTools xlsx export."""
    wb = load_workbook(filename=src, read_only=True, data_only=True)
    result = {}
    available_tabs = wb.sheetnames

    target_tabs = tabs if tabs else list(TAB_FIELD_MAP.keys())
    for tab_name in target_tabs:
        if tab_name in available_tabs:
            ws = wb[tab_name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = rows[0]
                data_rows = rows[1:]
                result[tab_name.lower()] = _parse_sheet(
                    headers, data_rows, tab_name
                )
            else:
                result[tab_name.lower()] = []
        else:
            result[tab_name.lower()] = []
    wb.close()
    return result


def _parse_csv(src, src_map, tabs, encoding):
    """Parse RVTools CSV export(s)."""
    result = {}
    target_tabs = tabs if tabs else list(TAB_FIELD_MAP.keys())

    sources = {}
    if src_map:
        sources = src_map
    elif src:
        tab_name = None
        for t in target_tabs:
            if t.lower() in os.path.basename(src).lower():
                tab_name = t
                break
        if tab_name:
            sources = {tab_name: src}

    for tab_name in target_tabs:
        csv_path = sources.get(tab_name)
        if csv_path and os.path.isfile(csv_path):
            with open(csv_path, encoding=encoding, newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)
                if rows:
                    headers = rows[0]
                    data_rows = rows[1:]
                    result[tab_name.lower()] = _parse_sheet(
                        headers, data_rows, tab_name
                    )
                else:
                    result[tab_name.lower()] = []
        else:
            result[tab_name.lower()] = []

    return result


def main():
    module = AnsibleModule(
        argument_spec=dict(
            src=dict(type="path", required=False),
            src_map=dict(type="dict", required=False),
            format=dict(type="str", required=True, choices=["xlsx", "csv"]),
            tabs=dict(type="list", elements="str", required=False, default=[]),
            encoding=dict(type="str", required=False, default="utf-8-sig"),
        ),
        required_if=[
            ("format", "xlsx", ("src",)),
        ],
        required_one_of=[
            ("src", "src_map"),
        ],
        supports_check_mode=True,
    )

    src = module.params["src"]
    src_map = module.params["src_map"]
    fmt = module.params["format"]
    tabs = module.params["tabs"]
    encoding = module.params["encoding"]

    if fmt == "xlsx":
        if not HAS_OPENPYXL:
            module.fail_json(
                msg="openpyxl is required for xlsx format: pip install openpyxl",
                exception=OPENPYXL_IMPORT_ERROR,
            )
        if not os.path.isfile(src):
            module.fail_json(msg=f"Source file not found: {src}")
        try:
            parsed = _parse_xlsx(src, tabs)
        except Exception as e:
            module.fail_json(msg=f"Failed to parse xlsx: {e}")
    else:
        if src and not src_map and not os.path.isfile(src):
            module.fail_json(msg=f"Source file not found: {src}")
        try:
            parsed = _parse_csv(src, src_map, tabs, encoding)
        except Exception as e:
            module.fail_json(msg=f"Failed to parse csv: {e}")

    tab_counts = {k: len(v) for k, v in parsed.items()}

    module.exit_json(changed=False, parsed=parsed, tab_counts=tab_counts)


if __name__ == "__main__":
    main()
